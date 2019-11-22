from openpyxl import load_workbook
from konlpy.tag import Kkma
from konlpy.tag import Twitter
from collections import Counter
import pandas as pd
import numpy as np
import openpyxl
twitter = Twitter()
from konlpy.utils import pprint

wb = load_workbook('search.xlsx')
sheet = wb['Sheet']

row = 1

sentences = list()

while True:
    text = sheet.cell(row=row, column=1).value

    if not text:
        break

    sentences.append(text)

    row += 1

kkma = Kkma()

poslist = list()

sentences_tag = []
for sentence2 in sentences:
    morph = twitter.pos(sentence2)
    sentences_tag.append(morph)


noun_adj_list = []
for sentence3 in sentences_tag:
    for word, tag in sentence3:
        if tag in ['Noun','Adjective']:
            noun_adj_list.append(word)

counts = Counter(noun_adj_list)
tag_count = []
tags = []
for n, c in counts.most_common(15):
    dics = {'tag':n, 'counts':c}
    if len(dics['tag']) >= 2 and len(tags) <= 49:
        tag_count.append(dics)
        tags.append(dics['tag'])
result = openpyxl.Workbook()
res_sheet = result.active
for tag in tag_count:
    print(" {:<14}".format(tag['tag']), end='\t')
    print("{}".format(tag['counts']))
    res_sheet.append([tag['tag'], tag['counts']])

result.save('daejeo_result.xlsx')


# print(counts.most_common(5))
# test_val = [counts.most_common(5)]
# data = pd.DataFrame.from_records(test_val)
# data.to_excel('result.xlsx')
